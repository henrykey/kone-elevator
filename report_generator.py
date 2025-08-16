# Author: IBC-AI CO.
"""
KONE 测试报告生成器
负责生成多种格式的测试报告：Markdown、JSON、HTML、Excel
"""

import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
from dataclasses import dataclass

# Optional Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional Jinja2 support - fallback to string templates
try:
    from jinja2 import Template
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False

logger = logging.getLogger(__name__)


@dataclass
class AuthTokenInfo:
    """Token和鉴权验证信息"""
    requested_scope: str  # 请求的scope
    token_scopes: str     # Token实际包含的scope
    is_match: bool        # 是否匹配
    error_message: Optional[str] = None  # 错误信息
    token_type: Optional[str] = None     # Token类型
    expires_in: Optional[int] = None     # 过期时间
    timestamp: Optional[str] = None      # 验证时间戳


@dataclass
@dataclass
class APICallInfo:
    """API调用详细信息"""
    interface_type: str  # 接口类型: "WebSocket" 或 "HTTP REST"
    url: str            # 请求URL
    method: Optional[str] = None  # HTTP方法 (GET/POST/PUT/DELETE) 或 WebSocket消息类型
    request_parameters: Optional[Dict] = None  # 发送的请求参数
    response_data: Optional[List[Dict]] = None  # 响应数据 (限制前1-2组)
    status_code: Optional[int] = None    # HTTP状态码
    timestamp: Optional[str] = None      # 调用时间戳
    error_message: Optional[str] = None  # 错误信息


@dataclass
class TestResult:
    """测试结果数据结构 - 符合KONE测试指南格式，增强版包含详细请求响应信息"""
    test_id: str  # Test编号 (如 "Test 1", "Test 2")
    name: str  # Test名称
    description: str  # Description描述
    expected_result: str  # Expected result期望结果
    test_result: str  # Test result (PASS/FAIL/TO_BE_FILLED)
    status: str  # 内部状态 (PASS, FAIL, SKIP, ERROR)
    duration_ms: float
    error_message: Optional[str] = None
    response_data: Optional[Dict] = None
    category: Optional[str] = None
    
    # 详细的API调用信息
    api_calls: Optional[List[APICallInfo]] = None  # 所有API调用的详细信息
    
    # 兼容性字段 (保留原有字段)
    request_parameters: Optional[Dict] = None  # 主要请求参数
    request_method: Optional[str] = None       # 主要HTTP方法
    request_url: Optional[str] = None          # 主要请求URL
    response_status_code: Optional[int] = None # 主要HTTP状态码
    response_headers: Optional[Dict] = None    # 响应头
    request_timestamp: Optional[str] = None    # 请求时间戳
    response_timestamp: Optional[str] = None   # 响应时间戳


class ReportGenerator:
    """
    KONE测试报告生成器
    支持生成符合《KONE Service Robot API Solution Validation Test Guide》格式的多种报告
    包含Setup、Pre-Test Setup、Date等元信息字段
    """
    
    def __init__(self, company_name: str = "IBC-AI CO."):
        """
        初始化报告生成器
        
        Args:
            company_name: 公司名称
        """
        self.company_name = company_name
        self.report_timestamp = datetime.now()
        self.auth_token_info: List[AuthTokenInfo] = []  # 存储Token验证信息
        
        # 从指南获取的测试用例映射
        self.test_guide_mapping = self._load_test_guide_mapping()
        logger.info(f"ReportGenerator initialized for {company_name}")
    
    def add_auth_token_info(self, auth_info: AuthTokenInfo):
        """添加Token验证信息"""
        self.auth_token_info.append(auth_info)
        logger.debug(f"Added auth token info: {auth_info.requested_scope} -> {auth_info.is_match}")
    
    def clear_auth_token_info(self):
        """清空Token验证信息"""
        self.auth_token_info.clear()
    
    def _generate_auth_section(self) -> str:
        """生成Token验证部分的Markdown内容"""
        if not self.auth_token_info:
            return """This section provides detailed information about OAuth2 token validation and scope verification. 
Token validation helps identify whether test failures are due to authentication/authorization issues rather than test script problems.

**No authentication data recorded for this session.**

### Expected Authentication Flow
1. **Token Request**: Client requests access token with specific scope
2. **Scope Validation**: Verify token contains required scope for operation  
3. **API Access**: Use validated token for WebSocket and REST API calls
4. **Error Handling**: Handle 401/403 errors with appropriate scope retry logic

**Recommendation**: Ensure authentication data is collected during test execution for better debugging capability."""
        
        auth_section = """This section provides detailed information about OAuth2 token validation and scope verification. 
Token validation helps identify whether test failures are due to authentication/authorization issues rather than test script problems.

### Token Scope Validation Results

| Requested Scope | Token Scopes | Match | Status | Error Message |
|----------------|--------------|-------|--------|---------------|
"""
        
        for auth_info in self.auth_token_info:
            status_icon = "✅" if auth_info.is_match else "❌"
            error_msg = auth_info.error_message or "-"
            
            auth_section += f"| `{auth_info.requested_scope}` | `{auth_info.token_scopes}` | {auth_info.is_match} | {status_icon} | {error_msg} |\n"
        
        # 添加统计信息
        total_auths = len(self.auth_token_info)
        successful_auths = len([auth for auth in self.auth_token_info if auth.is_match])
        failed_auths = total_auths - successful_auths
        
        auth_section += f"""
### Authentication Summary

- **Total Scope Validations**: {total_auths}
- **Successful Validations**: {successful_auths} ✅
- **Failed Validations**: {failed_auths} ❌
- **Success Rate**: {(successful_auths / total_auths * 100):.1f if total_auths > 0 else 0}%

### Common Authentication Issues

"""
        
        if failed_auths > 0:
            auth_section += """**⚠️ Authentication failures detected!**

Common causes of authentication failures:
- **Invalid Scope**: Token does not contain the required scope for the operation
- **Expired Token**: Token has expired and needs to be refreshed  
- **Wrong Building/Group**: Scope format incorrect for target building/group
- **API Credentials**: Invalid client_id or client_secret in configuration

**Troubleshooting Steps:**
1. Verify the scope format matches: `callv2/group:{buildingId}:{groupId}`
2. Check token expiration and refresh logic
3. Validate client credentials in config.yaml
4. Ensure building and group IDs are correct
"""
        else:
            auth_section += """**✅ All authentication validations successful!**

All requested scopes were properly validated and tokens contained the required permissions."""
        
        return auth_section
    
    def _load_test_guide_mapping(self) -> Dict[str, Dict[str, str]]:
        """加载测试指南中的测试用例映射"""
        return {
            # 1. Setup
            "Test_1": {
                "name": "Solution initialization",
                "description": "Solution initialization",
                "expected_result": "- Connections established by solution to test environment (Virtual or Preproduction).\n- Authentication successful\n- Get resources successful\n- Building config can be obtained.\n- Response code 200\n- Response code 401 in case if there is issue with API Credentials\n- Building actions can be obtained.\n- Response code 200\n- Response code 401 in case if there is issue with API Credentials",
                "category": "Setup"
            },
            
            # 2. Elevator mode check
            "Test_2": {
                "name": "Elevator non-operational mode",
                "description": "Is the elevator mode operational or not? Note: elevator mode is set to non-operational. Use if applicable to robot use case",
                "expected_result": "- Elevator mode is true with any of the below\n- Fire mode (FRD)\n- Out of service mode (OSS)\n- Attendant mode (ATS)\n- Priority mode (PRC)\n- call is not made",
                "category": "Elevator Mode Check"
            },
            "Test_3": {
                "name": "Elevator operational mode",
                "description": "Is the elevator mode operational or not? Note: elevator is set to operational; Source (any floor) – Destination (any floor)",
                "expected_result": "- Elevator mode is false with all below\n- Fire mode (FRD)\n- Out of service mode (OSS)\n- Attendant mode (ATS)\n- Priority mode (PRC)\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator destination is correct and as requested",
                "category": "Elevator Mode Check"
            },
            
            # 3. Elevator Call Giving (Tests 4-20)
            "Test_4": {
                "name": "Basic call",
                "description": "Call: Basic call -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking\n- Floor markings are as expected\n- Floor order is as expected\n- Elevator destination is correct as requested",
                "category": "Elevator Call Giving"
            },
            "Test_5": {
                "name": "Hold open elevator door",
                "description": "Call: hold open elevator door -> at Source floor, at Destination floor Note: Use if applicable to robot use case Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Elevator door stays open for\n- duration specified in hard time\n- optionally plus duration specified in soft time",
                "category": "Elevator Call Giving"
            },
            "Test_6": {
                "name": "Action call with invalid action id",
                "description": "Call: Action call with action id = 200, 0 [Unlisted action (range as in action payload)] -> Source: any floor, Destination: any floor. Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \" Ignoring call, unknown call action: {action id}\"\n- error message - \" Ignoring call, unknown call action: UNDEFINED\" if 0",
                "category": "Elevator Call Giving"
            },
            "Test_7": {
                "name": "Disabled action call",
                "description": "Call: Disabled action call with action id = 4 [listed enabled action (range as in action payload)] -> Source: any floor, Destination: any floor. Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \" Ignoring call, disabled call action: {{actionid}}\"",
                "category": "Elevator Call Giving"
            },
            "Test_8": {
                "name": "Mixed action call (invalid direction)",
                "description": "Call: Mixed action call with action id = 2002 -> Source: any floor. Note: This is applicable for Landing Call only. Elevator at first floor and direction down. use if applicable to robot use case",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \"INVALID_DIRECTION \"",
                "category": "Elevator Call Giving"
            },
            "Test_9": {
                "name": "Delay call (valid delay)",
                "description": "Call: Delay call with delay = 5 -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only. use if applicable to robot use case",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Elevator Call Giving"
            },
            "Test_10": {
                "name": "Delay call (invalid delay >30s)",
                "description": "Call: Delay call with delay = 40 [Invalid delay (range 0-30sec)] -> source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \" Invalid json payload\"",
                "category": "Elevator Call Giving"
            },
            "Test_11": {
                "name": "Transfer floor call",
                "description": "Call: Transfer floor call -> Source: any floor, Destination: any floor Note: The source and destination floors cannot be served by the same elevator. Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- modified destination included\n- modified reason included\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Elevator Call Giving"
            },
            "Test_12": {
                "name": "Through lift call (same floor, opposite sides)",
                "description": "Call: Through lift call -> Source: any floor, Destination: any floor Note: Both source and destination floors are on the same floor but opposite side of the elevator.",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- cancel Reason \"SAME_SOURCE_AND_DEST_FLOOR\"",
                "category": "Elevator Call Giving"
            },
            "Test_13": {
                "name": "No travel call (same floor)",
                "description": "Call: No travel call -> Source: any floor, Destination: same as source floor Note: Both source and destination floors are on the same floor and same side of the elevator.",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- cancel Reason \"SAME_SOURCE_AND_DEST_FLOOR\"",
                "category": "Elevator Call Giving"
            },
            "Test_14": {
                "name": "Specific lift call (allowed lifts)",
                "description": "Call: Specific lift Call-> Source: any floor, Destination: any floor Note: Allowed Lift \"allowed lifts\" id to be included in the send request. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator starts moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed",
                "category": "Elevator Call Giving"
            },
            "Test_15": {
                "name": "Cancel call",
                "description": "Call: Cancel Call-> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Request Elevator to move to any floor\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- Send request cancel with payload including Session id\n- Response code 201\n- Elevator stop moving",
                "category": "Elevator Call Giving"
            },
            "Test_16": {
                "name": "Null call (invalid destination)",
                "description": "Call: Null call -> Source: any floor, Destination: any floor Note: Destination floor is invalid and not part of building config. Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \"Ignoring call, unable to resolve destination: area:****\"",
                "category": "Elevator Call Giving"
            },
            "Test_17": {
                "name": "Null call (undefined destination)",
                "description": "Call: Null call -> Source: any floor, Destination: - Note: Destination floor is not defined. Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \" Ignoring call, destination not defined",
                "category": "Elevator Call Giving"
            },
            "Test_18": {
                "name": "Null call (invalid source)",
                "description": "Call: Null call -> Source: any floor, Destination: any floor Note: Source floor is invalid and not part of building config. Landing Call – Source only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \"Ignoring call, unable to resolve area: area:****\"",
                "category": "Elevator Call Giving"
            },
            "Test_19": {
                "name": "Null call (invalid source & destination)",
                "description": "Call: Null call -> Source: any floor, Destination: any floor Note: Source floor and Destination floor are both invalid and not part of building config.",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 201\n- error message - \"Ignoring call, unable to resolve area: area:****\"",
                "category": "Elevator Call Giving"
            },
            "Test_20": {
                "name": "Misplaced call (invalid building ID)",
                "description": "Call: Misplaced call to Building ID: a4KrX2cei -> Source: any floor, Destination: any floor Note: Building ID is invalid and not part of used resources. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Option 1: Illegal call prevented by robot controller\n- Option 2: Call allowed and Call cancelled\n- Response code 404\n- error message - \"Building data not found for ID building:a4KrX2cei\"",
                "category": "Elevator Call Giving"
            },
            
            # 4. Multiple groups and call giving
            "Test_21": {
                "name": "Group 2 access (second building id)",
                "description": "Group 2 (second building id) of physical building's lobby 2 access is provided. lobby 2 call -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Possible to select between groups (group 1 or group 2)\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Multiple Groups and Call Giving"
            },
            "Test_22": {
                "name": "Group 2 access (group suffix :2)",
                "description": "Group 2 (suffix :2) of physical building's lobby 2 access is provided. lobby 2 call -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Possible to select between groups (group 1 or group 2)\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Multiple Groups and Call Giving"
            },
            
            # 5. Integration with access control and call giving
            "Test_23": {
                "name": "Access control call (restricted floors)",
                "description": "Call: Access control call -> Source: any floor, Destination: any floor Note: floors are as defined in the access control permissions. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Robot have access only to the floors specified in the access rights\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Integration with Access Control and Call Giving"
            },
            "Test_24": {
                "name": "Access control call with media id",
                "description": "Call: Access control call with media id -> Source: any floor, Destination: any floor Note: floors are already defined in the access control system linked with the media id (with and without company code). Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Optional Robot have access only to the floors specified in the access rights\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Integration with Access Control and Call Giving"
            },
            
            # 6. Location control and call giving
            "Test_25": {
                "name": "Geographically controlled call",
                "description": "Call: Geographically controlled call -> Source: any floor, Destination: any floor Note: users are only allowed to place a call within the building they are located. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- While out of range, solution disables all calls\n- Calls cannot be made\n- While in range, solution enable all calls\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Location Control and Call Giving"
            },
            "Test_26": {
                "name": "Barrier call (before/after barrier)",
                "description": "Call: Barrier call -> Source: any floor, Destination: any floor Note: user is only allowed to place a call after crossing the barrier. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Before barrier, solution disables all calls\n- Calls cannot be made\n- After barrier, solution enables all calls\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Location Control and Call Giving"
            },
            "Test_27": {
                "name": "Barrier control and call (robot releases barrier)",
                "description": "Call: Barrier control and call -> Source: any floor, Destination: any floor Note: user is only allowed to place a call after crossing the barrier. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Before barrier, solution disables all calls\n- Calls cannot be made\n- At barrier, solution releases barrier for passage\n- After barrier, solution enables all calls\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Location Control and Call Giving"
            },
            "Test_28": {
                "name": "Multiple automatic calls prevention",
                "description": "Call: Multiple automatic calls prevention -> Source: any floor, Destination: any floor Note: user should only place single call for a journey. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door\n- No other calls generated while elevator approaches",
                "category": "Location Control and Call Giving"
            },
            "Test_29": {
                "name": "Multiple automatic calls prevention (group id)",
                "description": "Call: Multiple automatic call prevention for Group (Lobby) 2 -> Source: any floor, Destination: any floor Note: user should only place single call for a journey to the correct building id. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Possible to automatically select between groups (building id 1 or building id 2)\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door\n- No other calls generated while elevator approaches",
                "category": "Location Control and Call Giving"
            },
            "Test_30": {
                "name": "Multiple automatic calls prevention (group suffix)",
                "description": "Call: Multiple automatic call prevention for Group (Lobby) 2 -> Source: any floor, Destination: any floor Note: user should only place single call for a journey to the correct group suffix. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Possible to automatically select between groups (group suffix 1 or group suffix 2)\n- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door\n- No other calls generated while elevator approaches",
                "category": "Location Control and Call Giving"
            },
            
            # 7. Elevator locks and call giving
            "Test_31": {
                "name": "Enable locks (locked floor)",
                "description": "Call: Enable locks -> Source: any floor, Destination: any floor Note: in this use case source floor is locked by ACS solution. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and cancelled\n- Response code 201\n- cancel Reason - \"FLOOR_IS_LOCKED\"",
                "category": "Elevator Locks and Call Giving"
            },
            "Test_32": {
                "name": "Disable locks (unlocked floor)",
                "description": "Call: Disable locks -> Source: any floor, Destination: any floor Note: in this use case source floor is unlocked by ACS solution. Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Elevator Locks and Call Giving"
            },
            
            # 8. Device disabling and call giving
            "Test_33": {
                "name": "Allocation interrupted (all elevators disabled)",
                "description": "Call: Elevator's allocation interrupted (all elevators disabled) -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and cancelled\n- Response code 201\n- cancel Reason - \"NO_LIFT_AVAILABLE \"\n- Indicate call failure to user (such as timeout)",
                "category": "Device Disabling and Call Giving"
            },
            "Test_34": {
                "name": "Allocation resumed (all elevators enabled)",
                "description": "Call: Elevator's allocation interrupted (all elevators enabled) -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Device Disabling and Call Giving"
            },
            "Test_35": {
                "name": "Communication interrupted (DTU disconnected)",
                "description": "Call: End-to-end communication interrupted (DTU disconnected) -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call stuck on created\n- Response code 201\n- error: 1005\n- Indicate call failure to user (such as timeout)",
                "category": "Device Disabling and Call Giving"
            },
            "Test_36": {
                "name": "Call failure, communication interrupted (ping check)",
                "description": "Call: Call failure, communication interrupted -> Ping building or group Note: Start a ping sequence and stop pinging after positive response is obtained",
                "expected_result": "- Ping failed\n- Communication restored\n- Ping Successful",
                "category": "Device Disabling and Call Giving"
            },
            "Test_37": {
                "name": "Communication enabled (DTU connected)",
                "description": "Call: End-to-end communication enabled (DTU connected) -> Source: any floor, Destination: any floor Note: Landing Call – Source only, Car Call – Destination only",
                "expected_result": "- Call accepted and elevator moving\n- Response code 201\n- Session id returned\n- Elevator tracking as needed\n- if applicable hold open elevator door",
                "category": "Device Disabling and Call Giving"
            },
            "Test_38": {
                "name": "Custom case",
                "description": "Custom case",
                "expected_result": "To be filled",
                "category": "Other System Functionality Checks"
            }
        }
    
    def _extract_test_number(self, test_id: str) -> int:
        """从test_id中提取测试编号"""
        import re
        # 支持 "Test 1", "Test_1", "1" 等格式
        match = re.search(r'(\d+)', test_id)
        return int(match.group(1)) if match else 0
    
    def _get_test_info_from_guide(self, test_id: str) -> Dict[str, str]:
        """从测试指南获取测试信息"""
        # 规范化test_id格式：支持 "Test 1", "Test_1", "1" 等
        test_num = self._extract_test_number(test_id)
        normalized_test_id = f"Test_{test_num}"
        
        guide_info = self.test_guide_mapping.get(normalized_test_id, {})
        return {
            "name": guide_info.get("name", "Unknown Test"),
            "description": guide_info.get("description", "To be filled"),
            "expected_result": guide_info.get("expected_result", "Test should execute successfully and return expected results"),
            "category": guide_info.get("category", "Unknown")
        }
    
    def generate_report(self, test_results: List[TestResult], metadata: Dict[str, Any], output_dir: str = ".", config: Dict[str, Any] = None) -> Dict[str, str]:
        """
        生成符合KONE测试指南格式的多格式测试报告
        
        Args:
            test_results: 测试结果列表
            metadata: 测试元数据
            output_dir: 输出目录路径
            config: 配置信息（包含solution provider等）
            
        Returns:
            dict: 包含不同格式报告的字典
        """
        try:
            # 补充缺失的测试用例字段
            enhanced_test_results = self._enhance_test_results(test_results)
            
            # 补充报告元信息
            enhanced_metadata = self._enhance_metadata(metadata)
            
            # 统计测试结果
            stats = self._calculate_statistics(enhanced_test_results)
            
            # 生成报告数据
            report_data = {
                "metadata": enhanced_metadata,
                "statistics": stats,
                "test_results": enhanced_test_results,
                "generation_time": self.report_timestamp.isoformat(),
                "company": self.company_name,
                "config": config or {}
            }
            
            # 生成各种格式的报告 - 暂时只生成JSON格式
            reports = {
                "json": self._generate_json_report(report_data)
                # 其他格式暂时禁用
                # "markdown": self._generate_markdown_report(report_data),
                # "html": self._generate_html_report(report_data),
                # "excel": self._generate_excel_report(report_data, output_dir)
            }
            
            logger.info(f"Generated reports in {len(reports)} formats")
            return reports
            
        except Exception as e:
            logger.error(f"Failed to generate reports: {e}")
            return {
                "error": f"Report generation failed: {str(e)}"
            }
    
    def _enhance_test_results(self, test_results: List[TestResult]) -> List[TestResult]:
        """根据测试指南补充测试结果字段"""
        enhanced_results = []
        
        for result in test_results:
            # 从指南获取标准信息
            guide_info = self._get_test_info_from_guide(result.test_id)
            
            # 如果字段缺失，从指南补充
            if not hasattr(result, 'description') or not result.description:
                result.description = guide_info["description"]
            if not hasattr(result, 'expected_result') or not result.expected_result:
                result.expected_result = guide_info["expected_result"]
            if not hasattr(result, 'test_result') or not result.test_result:
                result.test_result = "PASS" if result.status == "PASS" else "FAIL" if result.status == "FAIL" else "TO_BE_FILLED"
            if not hasattr(result, 'category') or not result.category:
                result.category = guide_info["category"]
                
            enhanced_results.append(result)
        
        return enhanced_results
    
    def _enhance_metadata(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """补充报告元信息"""
        enhanced = metadata.copy()
        
        # 补充指南要求的字段
        enhanced.setdefault("setup", "Get access to the equipment for testing:\n- Virtual equipment, available in KONE API portal\n- Preproduction equipment, by contacting KONE API Support")
        enhanced.setdefault("pre_test_setup", "Test environments available for the correct KONE API organization.\nBuilding id can be retrieved (/resource endpoint).")
        enhanced.setdefault("date", datetime.now().strftime("%d.%m.%Y"))
        enhanced.setdefault("solution_provider", "IBC-AI CO.")
        enhanced.setdefault("tested_system", "KONE Elevator Control Service")
        enhanced.setdefault("kone_sr_api_version", "v2.0")
        
        return enhanced
    
    def _calculate_statistics(self, test_results: List[TestResult]) -> Dict[str, Any]:
        """
        计算测试统计信息
        
        Args:
            test_results: 测试结果列表
            
        Returns:
            dict: 统计信息
        """
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r.status == "PASS"])
        failed_tests = len([r for r in test_results if r.status == "FAIL"])
        error_tests = len([r for r in test_results if r.status == "ERROR"])
        skipped_tests = len([r for r in test_results if r.status == "SKIP"])
        
        total_duration = sum(r.duration_ms for r in test_results)
        avg_duration = total_duration / total_tests if total_tests > 0 else 0
        
        # 按分类统计
        category_stats = {}
        for result in test_results:
            category = result.category or "Unknown"
            if category not in category_stats:
                category_stats[category] = {"total": 0, "passed": 0, "failed": 0}
            
            category_stats[category]["total"] += 1
            if result.status == "PASS":
                category_stats[category]["passed"] += 1
            elif result.status in ["FAIL", "ERROR"]:
                category_stats[category]["failed"] += 1
        
        success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        return {
            "total_tests": total_tests,
            "passed_tests": passed_tests,
            "failed_tests": failed_tests,
            "error_tests": error_tests,
            "skipped_tests": skipped_tests,
            "success_rate": round(success_rate, 2),
            "total_duration_ms": total_duration,
            "average_duration_ms": round(avg_duration, 2),
            "category_breakdown": category_stats
        }
    
    def _get_solution_provider_info(self, config):
        """Get solution provider information from config or defaults"""
        provider_info = config.get('solution_provider', {})
        
        return {
            'company_name': provider_info.get('company_name', 'To be filled'),
            'company_address': provider_info.get('company_address', 'To be filled'),
            'contact_person_name': provider_info.get('contact_person_name', 'To be filled'),
            'contact_email': provider_info.get('contact_email', 'To be filled'),
            'contact_phone': provider_info.get('contact_phone', 'To be filled'),
            'tester': provider_info.get('tester', 'Automated Test System')
        }
    
    def _generate_markdown_report(self, report_data: Dict[str, Any]) -> str:
        """
        生成符合KONE测试指南格式的Markdown格式报告
        
        Args:
            report_data: 报告数据
            
        Returns:
            str: Markdown报告内容
        """
        # 直接使用指南格式的模板，不依赖Jinja2
        return self._generate_markdown_simple(report_data)
    
    def _generate_markdown_with_jinja(self, report_data: Dict[str, Any]) -> str:
        """使用Jinja2模板生成Markdown报告"""
        template_str = """# KONE Service Robot API Validation Test Report

**Generated by:** {{ company }}  
**Date:** {{ generation_time }}  
**Test Framework:** {{ metadata.test_framework or "N/A" }}  
**API Version:** {{ metadata.api_version or "N/A" }}

---

## Executive Summary

| Metric | Value |
|--------|-------|
| Total Tests | {{ statistics.total_tests }} |
| Passed | {{ statistics.passed_tests }} ✅ |
| Failed | {{ statistics.failed_tests }} ❌ |
| Errors | {{ statistics.error_tests }} ⚠️ |
| Skipped | {{ statistics.skipped_tests }} ⏭️ |
| Success Rate | {{ "%.1f" | format(statistics.success_rate) }}% |
| Total Duration | {{ "%.2f" | format(statistics.total_duration_ms / 1000) }} seconds |

---

## Test Results Summary

| Test ID | Test Name | Status | Duration (ms) | Error |
|---------|-----------|--------|---------------|-------|
{% for test in test_results %}
| {{ test.test_id }} | {{ test.name }} | {{ test.status }} | {{ "%.1f" | format(test.duration_ms) }} | {{ test.error_message or "-" }} |
{% endfor %}

---

## Detailed Test Results

{% for test in test_results %}
### {{ test.test_id }}: {{ test.name }}

- **Status:** {{ test.status }}
- **Category:** {{ test.category or "N/A" }}
- **Duration:** {{ "%.2f" | format(test.duration_ms) }} ms
{% if test.error_message %}
- **Error:** {{ test.error_message }}
{% endif %}
{% if test.response_data %}
- **Response Data:** 
```json
{{ test.response_data | tojson }}
```
{% endif %}

{% endfor %}

---

*Report generated by KONE Test Automation System v2.0*  
*{{ company }} - {{ generation_time }}*
"""
        
        if JINJA2_AVAILABLE:
            template = Template(template_str)
            return template.render(**report_data)
        else:
            # Fallback to simple string replacement
            return self._generate_markdown_simple(report_data)
    
    def _generate_markdown_simple(self, report_data: Dict[str, Any]) -> str:
        """使用简单字符串模板生成符合KONE测试指南格式的Markdown报告"""
        stats = report_data["statistics"]
        metadata = report_data["metadata"]
        config = report_data.get("config", {})
        
        # Get provider info from config
        provider_info = self._get_solution_provider_info(config)
        
        report = f"""# KONE Service Robot API Solution Validation Test Report

**SR-API (Service Robot API)**  
**Version**: 2.0  
**Author**: {report_data["company"]}

---

## Document Purpose

Ensuring the quality and security of a solution is every developer's responsibility. This document gives guidance on evaluating the readiness of those solutions that use Service Robot API.

---

## Test Information

### Setup
{metadata.get("setup", "To be filled")}

### Pre-Test Setup  
{metadata.get("pre_test_setup", "To be filled")}

### Date
| Test Date (dd.mm.yyyy) | Value |
|-------------------------|-------|
| Test execution date     | {metadata.get("date", datetime.now().strftime("%d.%m.%Y"))} |

### Solution Provider
| Item                 | Value |
|----------------------|-------|
| Company name         | {provider_info['company_name']} |
| Company address      | {provider_info['company_address']} |
| Contact person name  | {provider_info['contact_person_name']} |
| Email                | {provider_info['contact_email']} |
| Telephone number     | {provider_info['contact_phone']} |
| Tester               | {provider_info['tester']} |

### Tested System
| Item                     | Value |
|--------------------------|-------|
| System name              | {metadata.get("tested_system", "KONE Elevator Control Service")} |
| System version           | {metadata.get("system_version", "To be filled")} |
| Software name            | {metadata.get("software_name", "To be filled")} |
| Software version         | {metadata.get("software_version", "To be filled")} |
| KONE SR-API              | {metadata.get("kone_sr_api_version", "v2.0")} |
| KONE test assistant email | {metadata.get("kone_assistant_email", "To be filled")} |

---

## Test Summary

| Metric | Value |
|--------|-------|
| Total Tests | {stats["total_tests"]} |
| Passed | {stats["passed_tests"]} ✅ |
| Failed | {stats["failed_tests"]} ❌ |
| Errors | {stats["error_tests"]} ⚠️ |
| Skipped | {stats["skipped_tests"]} ⏭️ |
| Success Rate | {stats["success_rate"]}% |
| Total Duration | {stats["total_duration_ms"] / 1000:.2f} seconds |

---

## Authentication & Token Scope Validation

{self._generate_auth_section()}

---

## Service Robot API Solution Validation Test Results

| Test | Description | Expected result | Test result |
|------|-------------|-----------------|-------------|
"""
        
        # 按测试ID排序并添加详细测试结果表格
        def extract_test_number(test_id):
            """从测试ID中提取数字，支持 'Test 1' 和 'Test_1' 格式"""
            import re
            match = re.search(r'(\d+)', test_id)
            return int(match.group(1)) if match else 0
        
        sorted_results = sorted(report_data["test_results"], key=lambda x: extract_test_number(x.test_id))
        
        for result in sorted_results:
            # 格式化期望结果和测试结果，处理多行文本
            expected_formatted = result.expected_result.replace('\n', '<br>')
            test_result_formatted = result.test_result.replace('\n', '<br>') if hasattr(result, 'test_result') else "To be filled"
            
            report += f"| {result.test_id} | {result.description} | {expected_formatted} | {test_result_formatted} |\n"
        
        
        # 添加详细测试信息部分
        report += "\n---\n\n## Detailed Test Information\n\n"
        
        for result in sorted_results:
            report += f"### {result.test_id}: {result.name}\n"
            report += f"**Status:** {'✅ PASS' if result.status == 'PASS' else '❌ FAIL' if result.status == 'FAIL' else '⚠️ ERROR'}  \n"
            report += f"**Duration:** {result.duration_ms:.2f} ms\n"
            report += f"**Category:** {result.category or 'N/A'}\n\n"
            
            # 添加请求详情
            if hasattr(result, 'request_parameters') and result.request_parameters:
                report += "#### Request Details\n"
                if hasattr(result, 'request_method') and result.request_method:
                    report += f"- **Method:** {result.request_method}\n"
                if hasattr(result, 'request_url') and result.request_url:
                    report += f"- **URL:** `{result.request_url}`\n"
                if hasattr(result, 'request_timestamp') and result.request_timestamp:
                    report += f"- **Timestamp:** {result.request_timestamp}\n"
                report += "- **Parameters:**\n"
                import json
                report += "```json\n"
                report += json.dumps(result.request_parameters, indent=2, ensure_ascii=False)
                report += "\n```\n\n"
            
            # 添加响应详情
            if hasattr(result, 'response_data') and result.response_data:
                report += "#### Response Details\n"
                if hasattr(result, 'response_status_code') and result.response_status_code:
                    report += f"- **Status Code:** {result.response_status_code}\n"
                if hasattr(result, 'response_timestamp') and result.response_timestamp:
                    report += f"- **Response Time:** {result.response_timestamp}\n"
                report += "- **Response Data:**\n"
                report += "```json\n"
                report += json.dumps(result.response_data, indent=2, ensure_ascii=False)
                report += "\n```\n\n"
            
            # 添加错误信息（如果有）
            if hasattr(result, 'error_message') and result.error_message:
                report += "#### Error Information\n"
                report += f"```\n{result.error_message}\n```\n\n"
                
        # 添加建议
        report += "\n---\n\n## Test Analysis and Recommendations\n\n"
        
        if stats["success_rate"] >= 90:
            report += f"✅ **Excellent Performance**: The system demonstrates high reliability with {stats['success_rate']}% success rate.\n"
        elif stats["success_rate"] >= 70:
            report += "⚠️ **Good Performance**: The system shows acceptable performance but may need minor improvements.\n"
        else:
            report += f"❌ **Needs Improvement**: The system requires significant attention with {stats['success_rate']}% success rate.\n"
        
        report += "\n### Key Areas for Improvement:\n"
        for category, cat_stats in stats["category_breakdown"].items():
            if cat_stats["failed"] > 0:
                report += f"- **{category}**: {cat_stats['failed']} failed tests need attention\n"
        
        report += f"\n---\n\n*Report generated on {report_data['generation_time']} by {report_data['company']}*\n"
        
        return report
    
    def _generate_json_report(self, report_data: Dict[str, Any]) -> str:
        """
        生成符合KONE测试指南格式的JSON格式报告
        
        Args:
            report_data: 报告数据
            
        Returns:
            str: JSON报告内容
        """
        # 转换TestResult对象为符合指南格式的字典
        json_data = {
            "document_info": {
                "title": "KONE Service Robot API Solution Validation Test Report",
                "sr_api_version": "2.0",
                "author": report_data["company"],
                "generation_time": report_data["generation_time"],
                "report_version": "1.0"
            },
            "setup": {
                "setup_description": report_data["metadata"].get("setup", "To be filled"),
                "pre_test_setup": report_data["metadata"].get("pre_test_setup", "To be filled")
            },
            "test_information": {
                "date": report_data["metadata"].get("date", "To be filled"),
                "solution_provider": {
                    "company_name": report_data["metadata"].get("solution_provider", "IBC-AI CO."),
                    "company_address": report_data["metadata"].get("company_address", "To be filled"),
                    "contact_person_name": report_data["metadata"].get("contact_person", "To be filled"),
                    "email": report_data["metadata"].get("contact_email", "To be filled"),
                    "telephone_number": report_data["metadata"].get("contact_phone", "To be filled"),
                    "tester": report_data["metadata"].get("tester", "To be filled")
                },
                "tested_system": {
                    "system_name": report_data["metadata"].get("tested_system", "KONE Elevator Control Service"),
                    "system_version": report_data["metadata"].get("system_version", "To be filled"),
                    "software_name": report_data["metadata"].get("software_name", "To be filled"),
                    "software_version": report_data["metadata"].get("software_version", "To be filled"),
                    "kone_sr_api": report_data["metadata"].get("kone_sr_api_version", "v2.0"),
                    "kone_test_assistant_email": report_data["metadata"].get("kone_assistant_email", "To be filled")
                }
            },
            "test_summary": report_data["statistics"],
            "authentication_validation": {
                "total_validations": len(self.auth_token_info),
                "successful_validations": len([auth for auth in self.auth_token_info if auth.is_match]),
                "failed_validations": len([auth for auth in self.auth_token_info if not auth.is_match]),
                "validations": [
                    {
                        "requested_scope": auth.requested_scope,
                        "token_scopes": auth.token_scopes,
                        "is_match": auth.is_match,
                        "error_message": auth.error_message,
                        "token_type": auth.token_type,
                        "expires_in": auth.expires_in,
                        "timestamp": auth.timestamp
                    }
                    for auth in self.auth_token_info
                ]
            },
            "test_results": [
                {
                    "test": result.test_id,
                    "description": getattr(result, 'description', 'To be filled'),
                    "expected_result": getattr(result, 'expected_result', 'To be filled'),
                    "test_result": getattr(result, 'test_result', '待填写'),
                    "status": result.status,
                    "duration_ms": result.duration_ms,
                    "error_message": result.error_message,
                    "response_data": result.response_data,
                    "category": result.category,
                    
                    # 详细的API调用信息
                    "api_calls": [
                        {
                            "interface_type": api_call.interface_type,
                            "url": api_call.url,
                            "method": api_call.method,
                            "request_parameters": api_call.request_parameters,
                            "response_data": api_call.response_data[:2] if api_call.response_data and len(api_call.response_data) > 2 else api_call.response_data,  # 限制前1-2组
                            "status_code": api_call.status_code,
                            "timestamp": api_call.timestamp,
                            "error_message": api_call.error_message
                        }
                        for api_call in (result.api_calls or [])
                    ] if hasattr(result, 'api_calls') and result.api_calls else [],
                    
                    # 兼容性字段（保留原有字段）
                    "request_parameters": result.request_parameters,
                    "request_method": result.request_method,
                    "request_url": result.request_url,
                    "response_status_code": result.response_status_code,
                    "response_headers": result.response_headers,
                    "request_timestamp": result.request_timestamp,
                    "response_timestamp": result.response_timestamp
                }
                for result in sorted(report_data["test_results"], key=lambda x: self._extract_test_number(x.test_id))
            ]
        }
        
        return json.dumps(json_data, indent=2, ensure_ascii=False)
    
    def _generate_html_report(self, report_data: Dict[str, Any]) -> str:
        """
        生成HTML格式报告
        
        Args:
            report_data: 报告数据
            
        Returns:
            str: HTML报告内容
        """
        template_str = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>KONE API Validation Test Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .header {
            border-bottom: 3px solid #0066cc;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        .header h1 {
            color: #0066cc;
            margin: 0;
        }
        .meta-info {
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .stat-card.success { background: linear-gradient(135deg, #4CAF50, #45a049); }
        .stat-card.failure { background: linear-gradient(135deg, #f44336, #da190b); }
        .stat-card.warning { background: linear-gradient(135deg, #ff9800, #e68900); }
        
        .test-result {
            border: 1px solid #ddd;
            margin: 10px 0;
            border-radius: 5px;
            overflow: hidden;
        }
        .test-header {
            padding: 15px;
            font-weight: bold;
            cursor: pointer;
        }
        .test-header.pass { background: #d4edda; color: #155724; }
        .test-header.fail { background: #f8d7da; color: #721c24; }
        .test-header.error { background: #fff3cd; color: #856404; }
        .test-header.skip { background: #e2e3e5; color: #6c757d; }
        
        .test-details {
            padding: 15px;
            background: #f8f9fa;
            display: none;
        }
        .test-details.show { display: block; }
        
        pre {
            background: #f1f3f4;
            padding: 10px;
            border-radius: 4px;
            overflow-x: auto;
        }
        
        .progress-bar {
            width: 100%;
            background: #e0e0e0;
            border-radius: 25px;
            overflow: hidden;
            height: 20px;
            margin: 10px 0;
        }
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #4CAF50, #45a049);
            transition: width 0.3s ease;
        }
    </style>
    <script>
        function toggleDetails(testId) {
            const details = document.getElementById('details-' + testId);
            details.classList.toggle('show');
        }
    </script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏢 KONE Service Robot API Validation Test Report</h1>
            <p><strong>Generated by:</strong> {{ company }} | <strong>Date:</strong> {{ generation_time }}</p>
        </div>
        
        <div class="meta-info">
            <h3>📋 Test Environment</h3>
            <p><strong>API Version:</strong> {{ metadata.api_version }}</p>
            <p><strong>Test Framework:</strong> {{ metadata.test_framework }}</p>
            <p><strong>Building ID:</strong> {{ metadata.building_id or 'Unknown' }}</p>
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>{{ statistics.total_tests }}</h3>
                <p>Total Tests</p>
            </div>
            <div class="stat-card success">
                <h3>{{ statistics.passed_tests }}</h3>
                <p>Passed ✅</p>
            </div>
            <div class="stat-card failure">
                <h3>{{ statistics.failed_tests + statistics.error_tests }}</h3>
                <p>Failed ❌</p>
            </div>
            <div class="stat-card">
                <h3>{{ statistics.success_rate }}%</h3>
                <p>Success Rate</p>
            </div>
        </div>
        
        <div class="progress-bar">
            <div class="progress-fill" style="width: {{ statistics.success_rate }}%"></div>
        </div>
        
        <h3>📊 Test Results</h3>
        {% for result in test_results %}
        <div class="test-result">
            <div class="test-header {{ result.status.lower() }}" onclick="toggleDetails('{{ result.test_id }}')">
                {{ result.test_id }}: {{ result.name }}
                <span style="float: right;">
                    {% if result.status == "PASS" %}✅{% elif result.status == "FAIL" %}❌{% elif result.status == "ERROR" %}⚠️{% else %}⏭️{% endif %}
                    {{ result.status }}
                </span>
            </div>
            <div id="details-{{ result.test_id }}" class="test-details">
                <p><strong>Category:</strong> {{ result.category or 'Unknown' }}</p>
                <p><strong>Duration:</strong> {{ "%.2f" | format(result.duration_ms) }} ms</p>
                {% if result.error_message %}
                <p><strong>Error:</strong> {{ result.error_message }}</p>
                {% endif %}
                {% if result.response_data %}
                <p><strong>Response Data:</strong></p>
                <pre>{{ result.response_data | tojson }}</pre>
                {% endif %}
            </div>
        </div>
        {% endfor %}
        
        <div class="meta-info" style="margin-top: 30px;">
            <h3>💡 Recommendations</h3>
            {% if statistics.success_rate >= 90 %}
            <p>✅ <strong>Excellent Performance:</strong> The system demonstrates high reliability with {{ statistics.success_rate }}% success rate.</p>
            {% elif statistics.success_rate >= 70 %}
            <p>⚠️ <strong>Good Performance:</strong> The system shows acceptable performance but may need minor improvements.</p>
            {% else %}
            <p>❌ <strong>Needs Improvement:</strong> The system requires significant attention with {{ statistics.success_rate }}% success rate.</p>
            {% endif %}
        </div>
        
        <footer style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #666;">
            <p>Report generated on {{ generation_time }} by {{ company }}</p>
        </footer>
    </div>
</body>
</html>"""
        
        if JINJA2_AVAILABLE:
            template = Template(template_str)
            return template.render(**report_data)
        else:
            # Fallback: simple HTML without template rendering
            return f"""<!DOCTYPE html>
<html>
<head><title>KONE Test Report</title></head>
<body>
<h1>KONE Service Robot API Validation Test Report</h1>
<p><strong>Generated by:</strong> {report_data['company']}</p>
<p><strong>Date:</strong> {report_data['generation_time']}</p>
<p><strong>Note:</strong> Full HTML report requires Jinja2. Please install with: pip install jinja2</p>
</body>
</html>"""
    
    def _generate_excel_report(self, report_data: Dict[str, Any], output_dir: str = ".") -> str:
        """
        生成符合KONE测试指南格式的Excel格式报告
        
        Args:
            report_data: 报告数据
            output_dir: 输出目录路径
            
        Returns:
            str: Excel文件路径或错误信息
        """
        if not EXCEL_AVAILABLE:
            return "Excel generation not available - openpyxl package not installed"
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 主报告工作表 - 符合测试指南格式
            ws_main = wb.active
            ws_main.title = "Test Report"
            
            # 添加标题
            ws_main['A1'] = "KONE Service Robot API Solution Validation Test Report"
            ws_main['A1'].font = Font(bold=True, size=16)
            ws_main.merge_cells('A1:D1')
            
            # 添加文档信息
            current_row = 3
            ws_main[f'A{current_row}'] = "SR-API (Service Robot API)"
            ws_main[f'A{current_row+1}'] = "Version: 2.0"
            ws_main[f'A{current_row+2}'] = f"Author: {report_data['company']}"
            current_row += 5
            
            # 添加测试信息部分
            metadata = report_data["metadata"]
            
            # Setup信息
            ws_main[f'A{current_row}'] = "Setup"
            ws_main[f'A{current_row}'].font = header_font
            ws_main[f'A{current_row}'].fill = header_fill
            current_row += 1
            ws_main[f'A{current_row}'] = metadata.get("setup", "待填写")
            current_row += 3
            
            # Pre-Test Setup
            ws_main[f'A{current_row}'] = "Pre-Test Setup"
            ws_main[f'A{current_row}'].font = header_font
            ws_main[f'A{current_row}'].fill = header_fill
            current_row += 1
            ws_main[f'A{current_row}'] = metadata.get("pre_test_setup", "待填写")
            current_row += 3
            
            # Date
            ws_main[f'A{current_row}'] = "Date"
            ws_main[f'B{current_row}'] = "Value"
            ws_main[f'A{current_row}'].font = header_font
            ws_main[f'B{current_row}'].font = header_font
            ws_main[f'A{current_row}'].fill = header_fill
            ws_main[f'B{current_row}'].fill = header_fill
            current_row += 1
            ws_main[f'A{current_row}'] = "Test Date (dd.mm.yyyy)"
            ws_main[f'B{current_row}'] = metadata.get("date", datetime.now().strftime("%d.%m.%Y"))
            current_row += 3
            
            # Solution Provider
            ws_main[f'A{current_row}'] = "Solution Provider"
            ws_main[f'B{current_row}'] = "Value"
            ws_main[f'A{current_row}'].font = header_font
            ws_main[f'B{current_row}'].font = header_font
            ws_main[f'A{current_row}'].fill = header_fill
            ws_main[f'B{current_row}'].fill = header_fill
            current_row += 1
            
            provider_info = [
                ("Company name", metadata.get("solution_provider", "IBC-AI CO.")),
                ("Company address", metadata.get("company_address", "待填写")),
                ("Contact person name", metadata.get("contact_person", "待填写")),
                ("Email", metadata.get("contact_email", "待填写")),
                ("Telephone number", metadata.get("contact_phone", "待填写")),
                ("Tester", metadata.get("tester", "待填写"))
            ]
            
            for item, value in provider_info:
                ws_main[f'A{current_row}'] = item
                ws_main[f'B{current_row}'] = value
                current_row += 1
            current_row += 2
            
            # Tested System
            ws_main[f'A{current_row}'] = "Tested System"
            ws_main[f'B{current_row}'] = "Value"
            ws_main[f'A{current_row}'].font = header_font
            ws_main[f'B{current_row}'].font = header_font
            ws_main[f'A{current_row}'].fill = header_fill
            ws_main[f'B{current_row}'].fill = header_fill
            current_row += 1
            
            system_info = [
                ("System name", metadata.get("tested_system", "KONE Elevator Control Service")),
                ("System version", metadata.get("system_version", "待填写")),
                ("Software name", metadata.get("software_name", "待填写")),
                ("Software version", metadata.get("software_version", "待填写")),
                ("KONE SR-API", metadata.get("kone_sr_api_version", "v2.0")),
                ("KONE test assistant email", metadata.get("kone_assistant_email", "待填写"))
            ]
            
            for item, value in system_info:
                ws_main[f'A{current_row}'] = item
                ws_main[f'B{current_row}'] = value
                current_row += 1
            current_row += 3
            
            # 测试结果表格
            ws_main[f'A{current_row}'] = "Service Robot API Solution Validation Test Results"
            ws_main[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 2
            
            # 表头
            headers = ["Test", "Description", "Expected result", "Test result"]
            for col, header in enumerate(headers, start=1):
                cell = ws_main.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            current_row += 1
            
            # 添加测试结果数据 - 按测试ID排序
            sorted_results = sorted(report_data["test_results"], key=lambda x: self._extract_test_number(x.test_id))
            
            for result in sorted_results:
                ws_main.cell(row=current_row, column=1, value=result.test_id).border = border
                ws_main.cell(row=current_row, column=2, value=getattr(result, 'description', '待填写')).border = border
                ws_main.cell(row=current_row, column=3, value=getattr(result, 'expected_result', '待填写')).border = border
                ws_main.cell(row=current_row, column=4, value=getattr(result, 'test_result', '待填写')).border = border
                
                # 根据状态设置行颜色
                if result.status == "PASS":
                    fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif result.status == "FAIL":
                    fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                else:
                    fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                
                for col in range(1, 5):
                    ws_main.cell(row=current_row, column=col).fill = fill
                
                current_row += 1
            
            # 创建统计汇总工作表
            ws_summary = wb.create_sheet(title="Test Summary")
            
            # 统计信息
            stats = report_data["statistics"]
            summary_data = [
                ("Total Tests", stats["total_tests"]),
                ("Passed Tests", stats["passed_tests"]),
                ("Failed Tests", stats["failed_tests"]),
                ("Error Tests", stats["error_tests"]),
                ("Skipped Tests", stats["skipped_tests"]),
                ("Success Rate (%)", stats["success_rate"]),
                ("Total Duration (ms)", stats["total_duration_ms"]),
                ("Average Duration (ms)", stats["average_duration_ms"])
            ]
            
            ws_summary['A1'] = "Test Statistics"
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            for i, (key, value) in enumerate(summary_data, start=3):
                ws_summary[f'A{i}'] = key
                ws_summary[f'B{i}'] = value
            
            # 自动调整列宽
            for ws in [ws_main, ws_summary]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
            # 简化列宽处理 - 直接设置而不遍历列
            try:
                # 为主报告工作表设置列宽
                ws_main.column_dimensions['A'].width = 15  # Test列
                ws_main.column_dimensions['B'].width = 40  # Description列  
                ws_main.column_dimensions['C'].width = 50  # Expected result列
                ws_main.column_dimensions['D'].width = 20  # Test result列
                
                # 为统计工作表设置列宽
                ws_summary.column_dimensions['A'].width = 25  # 统计项目
                ws_summary.column_dimensions['B'].width = 15  # 数值
            except Exception as e:
                logger.warning(f"Column width adjustment failed: {e}")
            
            # 保存文件
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"KONE_Validation_Report_{timestamp}.xlsx"
            filepath = output_path / filename
            
            wb.save(str(filepath))
            
            logger.info(f"Excel report saved to: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return f"Excel generation failed: {str(e)}"
    
    def save_reports_to_files(self, reports: Dict[str, str], base_filename: str) -> Dict[str, str]:
        """
        将报告保存到文件
        
        Args:
            reports: 报告字典
            base_filename: 基础文件名
            
        Returns:
            dict: 保存的文件路径字典
        """
        saved_files = {}
        
        try:
            reports_dir = Path("./reports")
            reports_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 保存各种格式的报告
            if "markdown" in reports:
                md_filename = f"{base_filename}_{timestamp}.md"
                md_filepath = reports_dir / md_filename
                with open(md_filepath, 'w', encoding='utf-8') as f:
                    f.write(reports["markdown"])
                saved_files["markdown"] = str(md_filepath)
                logger.info(f"Markdown report saved: {md_filepath}")
            
            if "json" in reports:
                json_filename = f"{base_filename}_{timestamp}.json"
                json_filepath = reports_dir / json_filename
                with open(json_filepath, 'w', encoding='utf-8') as f:
                    f.write(reports["json"])
                saved_files["json"] = str(json_filepath)
                logger.info(f"JSON report saved: {json_filepath}")
            
            if "html" in reports:
                html_filename = f"{base_filename}_{timestamp}.html"
                html_filepath = reports_dir / html_filename
                with open(html_filepath, 'w', encoding='utf-8') as f:
                    f.write(reports["html"])
                saved_files["html"] = str(html_filepath)
                logger.info(f"HTML report saved: {html_filepath}")
            
            # Excel文件路径已在_generate_excel_report中处理
            if "excel" in reports and not reports["excel"].startswith("Excel generation"):
                saved_files["excel"] = reports["excel"]
            
            return saved_files
            
        except Exception as e:
            logger.error(f"Failed to save reports: {e}")
            return {"error": f"File saving failed: {str(e)}"}
